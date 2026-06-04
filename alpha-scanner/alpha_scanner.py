"""Alpha Stock Scanner v2.0 — S&P 500 momentum picks for 1-4 week holds."""
import pandas as pd, numpy as np, yfinance as yf, requests, time, warnings, json as _json
from datetime import datetime, timedelta
from pathlib import Path
warnings.filterwarnings('ignore')

CONFIG = {
    'universe': 'SP500', 'top_n': 10, 'lookback_days': 300,
    'weights': {'technical':0.30,'earnings':0.27,'institutional':0.18,'options_flow':0.12,'sentiment':0.13},
    'alpaca_api_key':'', 'alpaca_secret_key':'',
    'alpaca_data_url':'https://data.alpaca.markets',
    'output_excel':'alpha_scan_results.xlsx',
    'output_csv':'alpha_scan_results.csv',
    'output_json':'alpha_scan_results.json',
}
try:
    import config as _cfg
    for a,k in [('UNIVERSE','universe'),('TOP_N','top_n'),('WEIGHTS','weights'),
                ('ALPACA_API_KEY','alpaca_api_key'),('ALPACA_SECRET_KEY','alpaca_secret_key')]:
        if hasattr(_cfg,a): CONFIG[k]=getattr(_cfg,a)
except ImportError: pass

def _load_dotenv():
    for p in [Path('.env'), Path(__file__).parent/'.env']:
        if p.exists():
            for line in open(p):
                line=line.strip()
                if line and '=' in line and not line.startswith('#'):
                    k,_,v=line.partition('='); k=k.strip(); v=v.strip().strip('"').strip("'")
                    if k=='ALPACA_API_KEY' and v: CONFIG['alpaca_api_key']=v
                    elif k=='ALPACA_SECRET_KEY' and v: CONFIG['alpaca_secret_key']=v
            print(f'  Loaded API keys from {p}'); return
_load_dotenv()

SECTOR_ETFS={'Information Technology':'XLK','Health Care':'XLV','Financials':'XLF',
             'Consumer Discretionary':'XLY','Communication Services':'XLC','Industrials':'XLI',
             'Consumer Staples':'XLP','Energy':'XLE','Utilities':'XLU','Real Estate':'XLRE','Materials':'XLB'}

def get_sp500_universe():
    import io
    url='https://en.wikipedia.org/wiki/List_of_S%26P_500_companies'
    headers={'User-Agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'}
    resp=requests.get(url,headers=headers,timeout=15); resp.raise_for_status()
    df=pd.read_html(io.StringIO(resp.text))[0][['Symbol','Security','GICS Sector','GICS Sub-Industry']]
    df.columns=['ticker','name','sector','sub_industry']
    df['ticker']=df['ticker'].str.replace('.','-',regex=False); return df

def get_universe():
    cfg=CONFIG['universe']
    if isinstance(cfg,list): return pd.DataFrame({'ticker':cfg,'name':'','sector':'','sub_industry':''})
    if cfg=='SP500': return get_sp500_universe()
    raise ValueError(f'Unknown universe: {cfg}')

def fetch_price_data(tickers, lookback_days=300):
    period=f'{lookback_days+60}d'; data={}
    for i in range(0,len(tickers),50):
        batch=tickers[i:i+50]
        print(f'  Price data: {i+len(batch)}/{len(tickers)} ({(i+len(batch))/len(tickers)*100:.0f}%)',end='\r')
        try:
            raw=yf.download(' '.join(batch),period=period,group_by='ticker',auto_adjust=True,progress=False,threads=True)
            for t in batch:
                try:
                    df=raw[t].dropna(subset=['Close']) if len(batch)>1 else raw.dropna(subset=['Close'])
                    if len(df)>=60: data[t]=df
                except Exception: pass
        except Exception as e: print(f'\n  Batch warning: {e}')
        time.sleep(0.3)
    print(f'\n  Loaded {len(data)}/{len(tickers)} tickers'); return data

def calc_vcp_score(price_df):
    details={}
    try:
        weekly=price_df.resample('W').agg({'Open':'first','High':'max','Low':'min','Close':'last','Volume':'sum'}).dropna()
        if len(weekly)<10: return 50.0,{'vcp_note':'insufficient_data'}
        last_price=float(price_df['Close'].iloc[-1]); high_52w=float(price_df['Close'].tail(252).max())
        pct_off=(last_price/high_52w-1)*100; details['vcp_pct_off_52w_high']=round(pct_off,1)
        base_score=np.clip(1+pct_off/25,0,1)
        recent=weekly.tail(9).copy(); recent['range_pct']=(recent['High']-recent['Low'])/recent['Low']*100
        ranges=recent['range_pct'].values; volumes=recent['Volume'].values
        contractions=sum(1 for j in range(1,len(ranges)) if ranges[j]<ranges[j-1])
        details['vcp_contractions']=int(contractions)
        vol_trend=np.polyfit(range(len(volumes)),volumes,1)[0]
        vol_score=np.clip(-vol_trend/(volumes.mean()+1e-9)*10,0,1)
        final_range=float(ranges[-1]); details['vcp_final_range_pct']=round(final_range,2)
        tightness_score=np.clip(1-final_range/6,0,1)
        lows=recent['Low'].values; low_trend=np.polyfit(range(len(lows)),lows,1)[0]
        low_score=np.clip(low_trend/(lows.mean()+1e-9)*50,0,1)
        score=round(np.mean([base_score,contractions/5,vol_score,tightness_score,low_score])*100,1)
        details['is_vcp']=(pct_off>-25 and contractions>=3 and final_range<4)
    except Exception: score=50.0; details={'vcp_note':'error'}
    return score,details

def calc_technical_score(ticker,price_df,spy_df):
    details={}
    try:
        close=price_df['Close'].astype(float); volume=price_df['Volume'].astype(float); last=float(close.iloc[-1])
        ma50=close.rolling(50).mean(); ma200=close.rolling(200).mean()
        p50=(last/ma50.iloc[-1]-1)*100 if not pd.isna(ma50.iloc[-1]) else 0
        p200=(last/ma200.iloc[-1]-1)*100 if not pd.isna(ma200.iloc[-1]) else 0
        details['pct_vs_ma50']=round(p50,2); details['pct_vs_ma200']=round(p200,2)
        ret20=(last/close.iloc[-20]-1)*100
        spy_ret=(float(spy_df['Close'].iloc[-1])/float(spy_df['Close'].iloc[-20])-1)*100 if spy_df is not None and len(spy_df)>=20 else 0
        rs20=ret20-spy_ret; details['rs_vs_spy_20d']=round(rs20,2)
        delta=close.diff(); gain=delta.clip(lower=0).rolling(14).mean(); loss=(-delta.clip(upper=0)).rolling(14).mean()
        rsi=100-100/(1+gain/loss.replace(0,np.nan)); rv=float(rsi.iloc[-1]); details['rsi_14']=round(rv,1)
        vol10=volume.iloc[-10:].mean(); vol30=volume.iloc[-30:].mean()
        vexp=(vol10/vol30-1)*100 if vol30>0 else 0; details['vol_expansion_pct']=round(vexp,1)
        lk=min(252,len(close)); p52=(last/close.iloc[-lk:].max()-1)*100; details['pct_from_52w_high']=round(p52,2)
        mom5=(last/close.iloc[-5]-1)*100 if len(close)>=5 else 0; details['momentum_5d']=round(mom5,2)
        if 50<=rv<=65: rs=1.0
        elif 40<=rv<50: rs=0.7+(rv-40)/10*0.3
        elif 65<rv<=75: rs=1.0-(rv-65)/10*0.4
        else: rs=max(0.3-max(rv-75,0)/25*0.3,0)
        comps=[np.clip((p50+10)/20,0,1),np.clip((p200+20)/40,0,1),np.clip((rs20+8)/16,0,1),
               rs,np.clip((vexp+50)/100,0,1),np.clip((p52+25)/25,0,1),np.clip((mom5+3)/6,0,1)]
        base=np.mean(comps)*100
        vcp,vdet=calc_vcp_score(price_df); details.update(vdet)
        score=round(base*0.70+vcp*0.30,1)
    except Exception: score=50.0; details={}
    return score,details

def fetch_beat_streak(ticker):
    result={'beat_rate':None,'avg_beat_pct':None,'consecutive_beats':0}
    try:
        stock=yf.Ticker(ticker); qe=None
        for getter in [lambda: stock.earnings_history, lambda: stock.get_earnings_dates(limit=10)]:
            if qe is not None: break
            try:
                df=getter()
                if df is None or df.empty: continue
                df.columns=[c.strip() for c in df.columns]; col_map={}
                for c in df.columns:
                    cl=c.lower().replace(' ','').replace('_','')
                    if 'epsestimate' in cl or ('estimate' in cl and 'reported' not in cl): col_map['estimate']=c
                    elif 'reportedeps' in cl or 'actual' in cl or 'reported' in cl: col_map['actual']=c
                if 'estimate' in col_map and 'actual' in col_map:
                    df=df.rename(columns={col_map['estimate']:'Estimate',col_map['actual']:'Actual'})
                    df=df.dropna(subset=['Actual','Estimate']); df=df[df['Estimate']!=0].tail(8)
                    if len(df)>=2: qe=df
            except Exception: pass
        if qe is None: return result
        qe=qe.copy(); qe['beat']=qe['Actual']>qe['Estimate']
        qe['beat_pct']=(qe['Actual']-qe['Estimate'])/qe['Estimate'].abs()*100
        result['beat_rate']=round(float(qe['beat'].mean()),3)
        result['avg_beat_pct']=round(float(qe['beat_pct'].mean()),2)
        streak=0
        for b in qe.sort_index(ascending=False)['beat']:
            if b: streak+=1
            else: break
        result['consecutive_beats']=int(streak)
    except Exception: pass
    return result

def fetch_fundamentals(ticker):
    try:
        stock=yf.Ticker(ticker); info=stock.info
        r={'eps_growth_yoy':info.get('earningsGrowth'),'revenue_growth_yoy':info.get('revenueGrowth'),
           'target_mean_price':info.get('targetMeanPrice'),'recommendation':info.get('recommendationMean'),
           'current_price':info.get('currentPrice') or info.get('regularMarketPrice'),'earnings_date':None}
        try:
            cal=stock.calendar
            if cal is not None and not cal.empty and 'Earnings Date' in cal.index:
                r['earnings_date']=cal.loc['Earnings Date'].iloc[0]
        except Exception: pass
        return r
    except Exception: return {}

def calc_earnings_score(fund,beats,price):
    comps=[]; details={}
    ed=fund.get('earnings_date')
    if ed:
        try:
            ts=pd.Timestamp(ed).tz_localize(None) if getattr(ed,'tzinfo',None) else pd.Timestamp(ed)
            d=(ts-pd.Timestamp.now()).days; details['days_to_earnings']=int(d)
            comps.append(1.0 if 14<=d<=28 else 0.7 if (7<=d<14 or 28<d<=45) else 0.35 if 0<=d<7 else 0.25)
        except: comps.append(0.5)
    else: comps.append(0.5)
    eg=fund.get('eps_growth_yoy')
    if eg is not None: details['eps_growth_yoy']=f'{eg*100:.1f}%'; comps.append(np.clip((eg*100+20)/70,0,1))
    else: comps.append(0.5)
    rg=fund.get('revenue_growth_yoy')
    if rg is not None: details['rev_growth_yoy']=f'{rg*100:.1f}%'; comps.append(np.clip((rg*100+5)/30,0,1))
    else: comps.append(0.5)
    tgt=fund.get('target_mean_price'); pr=price or fund.get('current_price')
    if tgt and pr and pr>0:
        up=(tgt/pr-1)*100; details['analyst_upside_pct']=f'{up:.1f}%'; comps.append(np.clip((up+5)/35,0,1))
    else: comps.append(0.5)
    br=beats.get('beat_rate'); cs=beats.get('consecutive_beats',0); ab=beats.get('avg_beat_pct')
    if br is not None:
        details['beat_rate']=f'{br*100:.0f}%'; details['consecutive_beats']=int(cs)
        details['avg_beat_pct']=f'{ab:.1f}%' if ab else 'N/A'
        bc=br*0.4+np.clip(cs/4,0,1)*0.4+(np.clip((ab+10)/20,0,1)*0.2 if ab else 0)
        comps.append(bc)
    else: comps.append(0.5)
    ws=[1.0,1.0,1.0,1.0,1.5]; tw=sum(ws)
    return round(sum(c*w for c,w in zip(comps,ws))/tw*100,1),details

def fetch_institutional_data(ticker):
    r={'inst_pct_held':None,'short_pct_float':None,'recent_upgrades':0,'recent_downgrades':0}
    try:
        stock=yf.Ticker(ticker); info=stock.info
        r['inst_pct_held']=info.get('heldPercentInstitutions'); r['short_pct_float']=info.get('shortPercentOfFloat')
        try:
            recs=stock.recommendations
            if recs is not None and not recs.empty:
                cutoff=pd.Timestamp.now()-pd.Timedelta(days=30)
                if recs.index.tz: recs.index=recs.index.tz_localize(None)
                for _,row in recs[recs.index>=cutoff].iterrows():
                    a=str(row.get('Action','')).lower()
                    if 'up' in a: r['recent_upgrades']+=1
                    elif 'down' in a: r['recent_downgrades']+=1
        except Exception: pass
    except Exception: pass
    return r

def calc_institutional_score(d):
    comps=[]; details={}
    inst=d.get('inst_pct_held')
    if inst is not None: details['inst_pct_held']=f'{inst*100:.1f}%'; comps.append(np.clip(1-abs(inst-0.75)/0.40,0,1))
    else: comps.append(0.5)
    net=d.get('recent_upgrades',0)-d.get('recent_downgrades',0); details['net_upgrades_30d']=int(net)
    comps.append(np.clip((net+3)/6,0,1))
    short=d.get('short_pct_float')
    if short is not None: details['short_pct_float']=f'{short*100:.1f}%'; comps.append(np.clip(1-abs(short-0.12)/0.12,0,1))
    else: comps.append(0.5)
    return round(np.mean(comps)*100,1),details

def fetch_options_flow(ticker):
    r={'call_put_vol_ratio':None,'unusual_call_count':None,'iv_30d':None,'options_oi':None,'options_note':'pending'}
    try:
        stock=yf.Ticker(ticker); expiries=stock.options
        if not expiries: r['options_note']='no_expiries'; return r
        tgt=datetime.now()+timedelta(days=30)
        best=min(expiries,key=lambda e:abs((datetime.strptime(e,'%Y-%m-%d')-tgt).days))
        chain=stock.option_chain(best); calls=chain.calls.copy(); puts=chain.puts.copy()
        for df in [calls,puts]: df['volume']=df['volume'].fillna(0); df['openInterest']=df['openInterest'].fillna(0)
        oi=int(calls['openInterest'].sum()); r['options_oi']=oi
        if oi<200: r['options_note']='illiquid'; return r
        cv=calls['volume'].sum(); pv=puts['volume'].sum()
        if pv>0: r['call_put_vol_ratio']=round(cv/pv,2)
        liq=calls[calls['openInterest']>0]
        unc=liq[(liq['volume']>100)&(liq['volume']/liq['openInterest']>2.0)]
        r['unusual_call_count']=int(len(unc))
        try:
            price=float(getattr(stock.fast_info,'last_price',0) or 0)
            if price<=0: price=float(calls['strike'].median())
            atm=calls[(calls['strike']>=price*0.95)&(calls['strike']<=price*1.05)&(calls['impliedVolatility']>0)]
            if not atm.empty: r['iv_30d']=round(float(atm['impliedVolatility'].mean())*100,1)
        except Exception: pass
        r['options_note']='ok'
    except Exception: r['options_note']='error'
    return r

def calc_options_flow_score(d):
    note=d.get('options_note','')
    if note in ('no_expiries','illiquid','error'): return 50.0,{'options_note':note}
    comps=[]; details={}
    cp=d.get('call_put_vol_ratio')
    if cp is not None: details['call_put_ratio']=cp; comps.append(np.clip((cp-0.5)/1.5,0,1))
    else: comps.append(0.5)
    unc=d.get('unusual_call_count',0) or 0; details['unusual_calls']=int(unc)
    comps.append(np.clip(0.2+unc/5*0.8,0,1))
    iv=d.get('iv_30d')
    if iv is not None: details['iv_30d']=iv; comps.append(np.clip(1-(iv-20)/50,0,1))
    else: comps.append(0.5)
    oi=d.get('options_oi',0) or 0; details['options_oi']=oi; comps.append(np.clip(oi/15000,0,1))
    return round(np.mean(comps)*100,1),details

def _alpaca_headers():
    return {'APCA-API-KEY-ID':CONFIG.get('alpaca_api_key',''),'APCA-API-SECRET-KEY':CONFIG.get('alpaca_secret_key','')}

def fetch_news_count(ticker):
    if CONFIG.get('alpaca_api_key'):
        try:
            end=datetime.utcnow(); start=end-timedelta(days=7)
            url=(f"{CONFIG['alpaca_data_url']}/v1beta1/news?symbols={ticker}"
                 f"&start={start.strftime('%Y-%m-%dT%H:%M:%SZ')}&end={end.strftime('%Y-%m-%dT%H:%M:%SZ')}&limit=50")
            resp=requests.get(url,headers=_alpaca_headers(),timeout=6)
            if resp.status_code==200: return len(resp.json().get('news',[]))
        except Exception: pass
    try:
        news=yf.Ticker(ticker).news or []; cutoff=time.time()-7*86400; cnt=0
        for n in news:
            ts=n.get('providerPublishTime') or n.get('published') or 0
            if isinstance(ts,str):
                try: ts=int(datetime.fromisoformat(ts.replace('Z','')).timestamp())
                except: ts=0
            if ts>cutoff: cnt+=1
        return cnt
    except Exception: return 0

def calc_sentiment_score(ticker,sector,sector_scores):
    comps=[]; details={}
    nc=fetch_news_count(ticker); details['news_7d']=nc; comps.append(np.clip(nc/8,0,1))
    ss=sector_scores.get(sector,0.5); details['sector_rs_score']=round(ss*100,1); comps.append(ss)
    return round(np.mean(comps)*100,1),details

def calc_sector_strength():
    tickers=list(SECTOR_ETFS.values())+['SPY']; print('  Calculating sector strength...')
    try:
        raw=yf.download(' '.join(tickers),period='35d',group_by='ticker',auto_adjust=True,progress=False,threads=True)
        spy_ret=float(raw['SPY']['Close'].iloc[-1])/float(raw['SPY']['Close'].iloc[-20])-1
        scores={}
        for sec,etf in SECTOR_ETFS.items():
            try:
                ret=float(raw[etf]['Close'].iloc[-1])/float(raw[etf]['Close'].iloc[-20])-1
                scores[sec]=float(np.clip(((ret-spy_ret)*100+5)/10,0,1))
            except: scores[sec]=0.5
        return scores
    except: return {k:0.5 for k in SECTOR_ETFS}

def score_one_stock(ticker,price_data,spy_data,meta,sector_scores):
    if ticker not in price_data: return None
    df=price_data[ticker]; price=float(df['Close'].iloc[-1])
    ts,td=calc_technical_score(ticker,df,spy_data)
    fund=fetch_fundamentals(ticker); bs=fetch_beat_streak(ticker); es,ed=calc_earnings_score(fund,bs,price)
    inst=fetch_institutional_data(ticker); ins,ind=calc_institutional_score(inst)
    opts=fetch_options_flow(ticker); os,od=calc_options_flow_score(opts)
    sec=meta.get('sector',''); ss,sd=calc_sentiment_score(ticker,sec,sector_scores)
    w=CONFIG['weights']
    alpha=round(ts*w.get('technical',0)+es*w.get('earnings',0)+ins*w.get('institutional',0)+os*w.get('options_flow',0)+ss*w.get('sentiment',0),1)
    return {
        'ticker':ticker,'name':meta.get('name',''),'sector':sec,'sub_industry':meta.get('sub_industry',''),
        'price':round(price,2),'alpha_score':alpha,'technical_score':ts,'earnings_score':es,
        'institutional_score':ins,'options_flow_score':os,'sentiment_score':ss,
        'pct_vs_ma50':td.get('pct_vs_ma50'),'pct_vs_ma200':td.get('pct_vs_ma200'),
        'rs_vs_spy_20d':td.get('rs_vs_spy_20d'),'rsi_14':td.get('rsi_14'),
        'vol_expansion_pct':td.get('vol_expansion_pct'),'pct_from_52w_high':td.get('pct_from_52w_high'),
        'momentum_5d':td.get('momentum_5d'),'vcp_contractions':td.get('vcp_contractions'),
        'vcp_final_range_pct':td.get('vcp_final_range_pct'),'is_vcp':td.get('is_vcp',False),
        'days_to_earnings':ed.get('days_to_earnings'),'eps_growth_yoy':ed.get('eps_growth_yoy'),
        'rev_growth_yoy':ed.get('rev_growth_yoy'),'analyst_upside':ed.get('analyst_upside_pct'),
        'beat_rate':ed.get('beat_rate'),'consecutive_beats':ed.get('consecutive_beats'),
        'avg_beat_pct':ed.get('avg_beat_pct'),'inst_pct_held':ind.get('inst_pct_held'),
        'short_pct_float':ind.get('short_pct_float'),'net_upgrades_30d':ind.get('net_upgrades_30d'),
        'call_put_ratio':od.get('call_put_ratio'),'unusual_calls':od.get('unusual_calls'),
        'iv_30d':od.get('iv_30d'),'options_oi':od.get('options_oi'),
        'news_7d':sd.get('news_7d'),'sector_rs_score':sd.get('sector_rs_score'),
    }

def write_json(df,path):
    records=[]
    for i,row in df.head(CONFIG['top_n']).iterrows():
        records.append({'rank':i+1,'ticker':row.get('ticker',''),'name':row.get('name',''),
            'sector':row.get('sector',''),'price':row.get('price'),'alpha_score':row.get('alpha_score'),
            'technical':row.get('technical_score'),'earnings':row.get('earnings_score'),
            'institutional':row.get('institutional_score'),'options':row.get('options_flow_score'),
            'sentiment':row.get('sentiment_score'),'rsi':row.get('rsi_14'),
            'rs_spy':row.get('rs_vs_spy_20d'),'pct_52w':row.get('pct_from_52w_high'),
            'iv_30d':row.get('iv_30d'),'is_vcp':bool(row.get('is_vcp',False)),
            'beats':row.get('consecutive_beats'),'beat_rate':row.get('beat_rate'),
            'days_earnings':row.get('days_to_earnings'),'analyst_up':row.get('analyst_upside'),
            'call_put':row.get('call_put_ratio'),'news_7d':row.get('news_7d')})
    out={'scan_date':datetime.now().strftime('%Y-%m-%d %H:%M ET'),'universe':CONFIG['universe'],
         'top_n':CONFIG['top_n'],'stocks':records}
    with open(path,'w') as f: _json.dump(out,f,indent=2,default=str)
    print(f'  JSON saved  -> {path}')

def write_excel(df,path):
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill,Font,Alignment,Border,Side
    from openpyxl.formatting.rule import ColorScaleRule
    from openpyxl.utils import get_column_letter
    wb=Workbook()
    def fill(h): return PatternFill('solid',fgColor=h)
    hf=Font(name='Calibri',bold=True,color='FFFFFF',size=11)
    tf=Font(name='Calibri',bold=True,color='FFFFFF',size=14)
    bf=Font(name='Calibri',size=10)
    ca=Alignment(horizontal='center',vertical='center')
    la=Alignment(horizontal='left',vertical='center')
    bdr=Border(bottom=Side(style='thin',color='CCCCCC'),right=Side(style='thin',color='CCCCCC'))
    def write_sheet(ws,ddf,title,subtitle=''):
        ws.row_dimensions[1].height=30; nc=len(ddf.columns)
        ws.merge_cells(f'A1:{get_column_letter(nc)}1')
        c=ws['A1']; c.value=title; c.font=tf; c.fill=fill('1F3864'); c.alignment=ca
        hr=2
        if subtitle:
            ws.row_dimensions[2].height=18; ws.merge_cells(f'A2:{get_column_letter(nc)}2')
            s=ws['A2']; s.value=subtitle; s.font=Font(name='Calibri',italic=True,color='FFFFFF',size=10)
            s.fill=fill('2E75B6'); s.alignment=ca; hr=3
        ws.row_dimensions[hr].height=22
        for ci,col in enumerate(ddf.columns,1):
            c=ws.cell(row=hr,column=ci,value=col.replace('_',' ').title())
            c.font=hf; c.fill=fill('2E75B6'); c.alignment=ca; c.border=bdr
        for ri,row in enumerate(ddf.itertuples(index=False),hr+1):
            ws.row_dimensions[ri].height=17; rf=fill('F2F2F2') if ri%2==0 else fill('FFFFFF')
            for ci,val in enumerate(row,1):
                c=ws.cell(row=ri,column=ci,value=val); c.font=bf; c.fill=rf; c.border=bdr; c.alignment=la
        for ci,col in enumerate(ddf.columns,1):
            w=min(max(ddf[col].astype(str).map(len).max(),len(col.replace('_',' ').title()))+2,36)
            ws.column_dimensions[get_column_letter(ci)].width=w
        return hr
    rd=datetime.now().strftime('%Y-%m-%d %H:%M')
    ws1=wb.active; ws1.title='Top Alpha Picks'; ws1.sheet_view.showGridLines=False
    top=df.head(CONFIG['top_n']).reset_index(drop=True).copy(); top.insert(0,'Rank',range(1,len(top)+1))
    cols=['Rank','ticker','name','sector','price','alpha_score','technical_score','earnings_score',
          'institutional_score','options_flow_score','sentiment_score','is_vcp','vcp_contractions',
          'rsi_14','rs_vs_spy_20d','pct_from_52w_high','days_to_earnings','beat_rate',
          'consecutive_beats','analyst_upside','net_upgrades_30d','call_put_ratio',
          'unusual_calls','iv_30d','options_oi','news_7d']
    avail=[c for c in cols if c in top.columns]
    hr=write_sheet(ws1,top[avail],f'Alpha Stock Scanner v2.0 -- Top {CONFIG["top_n"]} Candidates',
                   f'Run: {rd}  |  Universe: {CONFIG["universe"]}  |  Hold: 1-4 Weeks')
    ai=avail.index('alpha_score')+1; al=get_column_letter(ai)
    ws1.conditional_formatting.add(f'{al}{hr+1}:{al}{hr+len(top)}',
        ColorScaleRule(start_type='num',start_value=40,start_color='FF6B6B',
                       mid_type='num',mid_value=60,mid_color='FFD700',
                       end_type='num',end_value=80,end_color='70AD47'))
    ws1.freeze_panes=f'A{hr+1}'
    ws2=wb.create_sheet('Full Universe'); ws2.sheet_view.showGridLines=False
    full=df.reset_index(drop=True).copy(); full.insert(0,'Rank',range(1,len(full)+1))
    write_sheet(ws2,full,f'Full Universe -- {CONFIG["universe"]}',f'Sorted by Alpha Score | {rd}')
    ws2.freeze_panes='A4'
    wb.save(path); print(f'  Excel saved -> {path}')

def run_scanner():
    print('\n'+'='*64)
    print('  ALPHA STOCK SCANNER  v2.0')
    print(f'  {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}')
    print(f'  Options flow: ENABLED (yfinance)')
    print(f'  Alpaca news:  {"ENABLED" if CONFIG.get("alpaca_api_key") else "DISABLED"}')
    print('='*64)
    print('\n[1/5] Loading universe...')
    univ=get_universe(); tickers=univ['ticker'].tolist()
    meta={r['ticker']:r for r in univ.to_dict('records')}; print(f'  {len(tickers)} tickers')
    print('\n[2/5] Fetching price data...')
    price_data=fetch_price_data(tickers,CONFIG['lookback_days'])
    spy_data=yf.download('SPY',period='400d',auto_adjust=True,progress=False)
    print('\n[3/5] Sector relative strength...')
    sector_scores=calc_sector_strength()
    for sec,v in sorted(sector_scores.items(),key=lambda x:-x[1]):
        print(f'  {sec:<30} {v*100:5.1f}  {"#"*int(v*20)}')
    print(f'\n[4/5] Scoring stocks (15-25 min)...')
    results=[]; valid=list(price_data.keys())
    for i,ticker in enumerate(valid):
        if i%20==0: print(f'  {i}/{len(valid)} ({i/len(valid)*100:.0f}%)  -- {ticker:<8}',end='\r')
        try:
            r=score_one_stock(ticker,price_data,spy_data,meta.get(ticker,{}),sector_scores)
            if r: results.append(r)
        except Exception: pass
        time.sleep(0.07)
    print(f'\n  Scored {len(results)} stocks')
    print('\n[5/5] Ranking and saving...')
    df=(pd.DataFrame(results).sort_values('alpha_score',ascending=False).reset_index(drop=True))
    df.to_csv(CONFIG['output_csv'],index=False)
    write_excel(df,CONFIG['output_excel'])
    write_json(df,CONFIG.get('output_json','alpha_scan_results.json'))
    top_n=CONFIG['top_n']
    print(f'\n{"="*72}')
    print(f'  TOP {top_n} ALPHA CANDIDATES  --  {datetime.now().strftime("%Y-%m-%d")}')
    print(f'{"="*72}')
    print(f'  {"Rank":<4} {"Tick":<6} {"Name":<26} {"Alpha":>6} {"Tech":>5} {"Earn":>5} {"Inst":>5} {"Opts":>5} {"Sent":>5} {"IV%":>5} {"VCP":>4} {"Beats":>6}')
    print('  '+'-'*70)
    for i,row in df.head(top_n).iterrows():
        name=str(row.get('name',''))[:25]; vcp='VCP' if row.get('is_vcp') else ''
        beats=str(row.get('consecutive_beats','')) or ''; iv=f'{row["iv_30d"]:.0f}' if row.get('iv_30d') else '--'
        print(f'  {i+1:<4} {row["ticker"]:<6} {name:<26} {row["alpha_score"]:>6.1f} {row["technical_score"]:>5.1f} {row["earnings_score"]:>5.1f} {row["institutional_score"]:>5.1f} {row["options_flow_score"]:>5.1f} {row["sentiment_score"]:>5.1f} {iv:>5} {vcp:>4} {beats:>6}')
    print(f'\n  Excel -> {CONFIG["output_excel"]}')
    print(f'  CSV   -> {CONFIG["output_csv"]}')
    print(f'  JSON  -> {CONFIG.get("output_json","alpha_scan_results.json")}')
    return df

if __name__=='__main__':
    run_scanner()
